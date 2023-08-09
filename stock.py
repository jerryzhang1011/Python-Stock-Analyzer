from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from os.path import exists

def write_into_xl(input_file_path, xl_path):
    with open(input_file_path, "r") as file:
        rows = file.readlines()
    new_rows = []
    new_rows.append(["time", rows.pop(0)])
    categories = ["market index", "most active", "most gainers", "most losers"]
    col_types = ["rank", "stock name", "curr price", "change percentage today", "changed price today"]
    category = 0
    for i in rows:
        if i[:3] == " 1?":
            if category:
                new_rows.append([])
            new_rows.append(["category", categories[category]])
            new_rows.append(col_types)
            category += 1
        new_rows.append(i.split("?".strip()))
    new_rows.append([])

    if exists(xl_path):
        wb = load_workbook(xl_path)
    else:
        wb = Workbook()
    ws = wb.active
    ws.insert_rows(1, len(new_rows))
    for row_index, row_data in enumerate(new_rows, start=1):
        for col_index, value in enumerate(row_data, start=1):
            ws.cell(row=row_index, column=col_index, value=value)
    color = "000000"
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for col in range(1, 6):
        cell = ws.cell(row=len(new_rows) + 1, column=col)
        cell.fill = fill
    wb.save(xl_path)

write_into_xl("stocks/data.txt", "stocks/output.xlsx")
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Alignment, Font, numbers, NamedStyle
from os.path import exists


def write_into_xl(input_file_path, xl_path):
    # reading, interpreting, and saving the text file
    with open(input_file_path, "r") as file:
        rows = file.readlines()
    new_rows = []
    new_rows.append(["Time", rows.pop(0)])
    categories = ["Market Index", "Most Active", "Most Gainers", "Most Losers"]
    col_types = [
        "Rank",
        "Company",
        "Price (USD)",
        "Change ($)",
        "Change (%)",
    ]
    category = 0
    category_locations = []
    for i in rows:
        if i[:3] == " 1?":
            if category:
                new_rows.append([])
            new_rows.append(["Category", categories[category]])
            category_locations.append(len(new_rows))
            new_rows.append(col_types)
            category += 1
        new_rows.append(i.strip().replace(" ", "").split("?"))
    for i in range(3):
        new_rows.append([])

    # writing data into the excel file
    if exists(xl_path):
        wb = load_workbook(xl_path)
    else:
        wb = Workbook()
    ws = wb.active
    ws.insert_rows(1, len(new_rows))
    for row_index, row_data in enumerate(new_rows, start=1):
        for col_index, value in enumerate(row_data, start=1):
            ws.cell(row=row_index, column=col_index, value=value)

    # styles
    max_row = len(new_rows)
    max_col = 5
    # black line
    color = "000000"
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for col in range(1, max_col + 1):
        cell = ws.cell(row=max_row, column=col)
        cell.fill = fill

    # text alignment and font for all cells
    alignment = Alignment(horizontal="center", vertical="center")
    font = Font(name="Times New Roman")
    for row in ws.iter_rows(max_row=max_row, max_col=max_col):
        for cell in row:
            cell.alignment = alignment
            cell.font = font
    cell = ws["B1"]
    value = cell.value.split(" ", 1)
    cell.value = "\n".join(value)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

    # number formats
    formats = {
        1: numbers.FORMAT_NUMBER,
        3: numbers.FORMAT_NUMBER_00,
        4: numbers.FORMAT_CURRENCY_USD_SIMPLE,
        5: numbers.FORMAT_PERCENTAGE_00,
    }
    for col in [1, 3, 4, 5]:
        for row in range(1, max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None and any(char.isdigit() for char in cell.value):
                if col == 1:
                    cell.value = int(cell.value)
                elif col == 5:
                    cell.value = float(cell.value.replace("%", "")) / 100
                else:
                    cell.value = float(cell.value.replace("$", "").replace(",", ""))
                cell.number_format = formats[col]

    # heights
    ws.row_dimensions[1].height = 63
    for row in range(2, max_row + 1):
        if row in category_locations:
            ws.row_dimensions[row].height = 52
        elif row - 1 in category_locations:
            ws.row_dimensions[row].height = 27
        else:
            ws.row_dimensions[row].height = 20

    # widths
    widths = {"A": 10, "B": 31, "C": 15, "D": 15, "E": 15}
    for col in ["A", "B", "C", "D", "E"]:
        ws.column_dimensions[col].width = widths[col]

    # fonts for titles
    cell = ws["A1"]
    cell.font = Font(bold=True, size=16)
    cell = ws["B1"]
    cell.font = Font(bold=True, size=14)
    for row in category_locations:
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True, size=16)
            cell = ws.cell(row=row + 1, column=col)
            cell.font = Font(bold=True, size=14)
    wb.save(xl_path)